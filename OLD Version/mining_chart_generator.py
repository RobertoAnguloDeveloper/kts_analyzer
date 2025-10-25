import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Set matplotlib style
plt.style.use('seaborn-v0_8-darkgrid')

class MiningDataChartGenerator:
    def __init__(self, input_file, sheet_name=None):
        """Initialize the chart generator"""
        self.input_file = input_file
        self.sheet_name = sheet_name
        self.df = None
        self.clean_df = None
        self.temp_dir = "temp_charts"
        
        # Create temp directory for charts
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
            
    def load_data(self):
        """Load Excel data"""
        print(f"📂 Loading file: {self.input_file}")
        
        if self.sheet_name:
            self.df = pd.read_excel(self.input_file, sheet_name=self.sheet_name)
        else:
            # Get first sheet
            excel_file = pd.ExcelFile(self.input_file)
            sheet_names = excel_file.sheet_names
            print(f"   Available sheets: {', '.join(sheet_names)}")
            self.df = pd.read_excel(self.input_file, sheet_name=sheet_names[0])
            print(f"   Using sheet: {sheet_names[0]}")
            
        print(f"✅ Data loaded: {self.df.shape[0]} rows × {self.df.shape[1]} columns")
        
    def process_data(self):
        """Process and clean the data"""
        print("\n🔧 Processing data...")
        
        try:
            # Find header row containing date columns
            header_row = 0
            for idx in range(min(10, len(self.df))):
                row_values = self.df.iloc[idx].astype(str).values
                if any('ene-' in val.lower() or 'feb-' in val.lower() or 'mar-' in val.lower() 
                      for val in row_values):
                    header_row = idx
                    break
            
            # Set proper column names
            if header_row > 0:
                self.df.columns = self.df.iloc[header_row]
                self.df = self.df.iloc[header_row + 1:].reset_index(drop=True)
            
            # Clean column names
            self.df.columns = [str(col).strip() if pd.notna(col) else f'Col_{i}' 
                              for i, col in enumerate(self.df.columns)]
            
            print(f"   Found {len([c for c in self.df.columns if '-' in str(c)])} date columns")
            
            # Extract data into dictionary
            data_dict = {}
            date_columns = []
            
            for col in self.df.columns[3:]:  # Skip first 3 columns (metric info)
                if '-' in str(col):
                    date_columns.append(str(col))
            
            # Process each metric row
            for idx, row in self.df.iterrows():
                if pd.notna(row.iloc[0]):
                    # Get metric name
                    metric_name = str(row.iloc[0]).strip()
                    
                    # Get subcategory if exists
                    if pd.notna(row.iloc[1]):
                        subcategory = str(row.iloc[1]).strip()
                        if subcategory.lower() not in ['nan', '']:
                            key = f"{metric_name}_{subcategory}"
                        else:
                            key = metric_name
                    else:
                        key = metric_name
                    
                    # Clean key name
                    key = key.replace(' ', '_').replace('(', '').replace(')', '')
                    
                    # Extract values
                    values = []
                    for date_col in date_columns:
                        if date_col in self.df.columns:
                            val = row[date_col]
                            if pd.notna(val):
                                # Clean numeric values
                                if isinstance(val, str):
                                    val = val.replace(',', '').replace('.', '')
                                try:
                                    values.append(float(val))
                                except:
                                    values.append(0)
                            else:
                                values.append(0)
                    
                    if values and not all(v == 0 for v in values):
                        data_dict[key] = values
            
            # Create clean dataframe
            self.clean_df = pd.DataFrame(data_dict)
            
            # Parse dates
            parsed_dates = self._parse_dates(date_columns[:len(self.clean_df)])
            self.clean_df['Date'] = parsed_dates
            
            # Set date as index
            self.clean_df['Date'] = pd.to_datetime(self.clean_df['Date'], errors='coerce')
            self.clean_df = self.clean_df.dropna(subset=['Date'])
            self.clean_df = self.clean_df.set_index('Date').sort_index()
            
            print(f"✅ Processed {len(self.clean_df)} time periods with {len(self.clean_df.columns)} metrics")
            print(f"   Date range: {self.clean_df.index[0].strftime('%b %Y')} to {self.clean_df.index[-1].strftime('%b %Y')}")
            print(f"   Metrics found: {', '.join(list(self.clean_df.columns)[:5])}...")
            
        except Exception as e:
            print(f"❌ Error processing data: {str(e)}")
            raise
            
    def _parse_dates(self, date_strings):
        """Parse Spanish month abbreviations to dates"""
        months = {
            'ene': '01', 'jan': '01', 'enero': '01',
            'feb': '02', 'febrero': '02',
            'mar': '03', 'marzo': '03',
            'abr': '04', 'apr': '04', 'abril': '04',
            'may': '05', 'mayo': '05',
            'jun': '06', 'junio': '06',
            'jul': '07', 'julio': '07',
            'ago': '08', 'aug': '08', 'agosto': '08',
            'sep': '09', 'sept': '09', 'septiembre': '09',
            'oct': '10', 'octubre': '10',
            'nov': '11', 'noviembre': '11',
            'dic': '12', 'dec': '12', 'diciembre': '12'
        }
        
        parsed = []
        for date_str in date_strings:
            try:
                parts = str(date_str).lower().split('-')
                if len(parts) >= 2:
                    month = months.get(parts[0], '01')
                    year = parts[1]
                    if len(year) == 2:
                        year = '20' + year
                    parsed.append(f"{year}-{month}-01")
                else:
                    parsed.append(None)
            except:
                parsed.append(None)
        return parsed
    
    def generate_charts(self):
        """Generate all charts as images"""
        print("\n📊 Generating charts...")
        
        chart_files = []
        
        # 1. Production Overview
        print("   Creating Production Overview...")
        file1 = self._create_production_overview()
        if file1:
            chart_files.append(('Production Overview', file1))
        
        # 2. Efficiency Metrics
        print("   Creating Efficiency Analysis...")
        file2 = self._create_efficiency_charts()
        if file2:
            chart_files.append(('Efficiency Analysis', file2))
        
        # 3. Comparative Analysis
        print("   Creating Comparative Analysis...")
        file3 = self._create_comparative_charts()
        if file3:
            chart_files.append(('Comparative Analysis', file3))
        
        # 4. Trend Analysis
        print("   Creating Trend Analysis...")
        file4 = self._create_trend_charts()
        if file4:
            chart_files.append(('Trend Analysis', file4))
        
        print(f"✅ Generated {len(chart_files)} chart sets")
        return chart_files
    
    def _create_production_overview(self):
        """Create production overview charts"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Production Overview', fontsize=16, fontweight='bold')
        
        # Chart 1: Ore Mined Comparison
        ax1 = axes[0, 0]
        plotted = False
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            ax1.plot(self.clean_df.index, self.clean_df['Ore_Mined_RGM'], 
                    marker='o', label='RGM', linewidth=2, markersize=4)
            plotted = True
        if 'Ore_Mined_Sar' in self.clean_df.columns:
            ax1.plot(self.clean_df.index, self.clean_df['Ore_Mined_Sar'], 
                    marker='s', label='Sar', linewidth=2, markersize=4)
            plotted = True
        
        if plotted:
            ax1.set_title('Ore Production Over Time', fontweight='bold')
            ax1.set_xlabel('Date')
            ax1.set_ylabel('Ore Mined (kt)')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 2: Overburden Comparison
        ax2 = axes[0, 1]
        plotted = False
        if 'Overburden_RGM' in self.clean_df.columns:
            ax2.plot(self.clean_df.index, self.clean_df['Overburden_RGM'], 
                    marker='o', label='RGM', linewidth=2, markersize=4, color='orange')
            plotted = True
        if 'Overburden_Sar' in self.clean_df.columns:
            ax2.plot(self.clean_df.index, self.clean_df['Overburden_Sar'], 
                    marker='s', label='Sar', linewidth=2, markersize=4, color='green')
            plotted = True
            
        if plotted:
            ax2.set_title('Overburden Movement', fontweight='bold')
            ax2.set_xlabel('Date')
            ax2.set_ylabel('Overburden (kt)')
            ax2.legend()
            ax2.grid(True, alpha=0.3)
            ax2.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 3: Total Material Movement
        ax3 = axes[1, 0]
        total_ore = pd.Series(0, index=self.clean_df.index)
        total_overburden = pd.Series(0, index=self.clean_df.index)
        
        for col in ['Ore_Mined_RGM', 'Ore_Mined_Sar']:
            if col in self.clean_df.columns:
                total_ore += self.clean_df[col]
                
        for col in ['Overburden_RGM', 'Overburden_Sar']:
            if col in self.clean_df.columns:
                total_overburden += self.clean_df[col]
        
        if total_ore.sum() > 0 or total_overburden.sum() > 0:
            ax3.bar(self.clean_df.index, total_ore, label='Total Ore', alpha=0.7, color='skyblue')
            ax3.bar(self.clean_df.index, total_overburden, bottom=total_ore, 
                   label='Total Overburden', alpha=0.7, color='coral')
            ax3.set_title('Total Material Movement', fontweight='bold')
            ax3.set_xlabel('Date')
            ax3.set_ylabel('Material (kt)')
            ax3.legend()
            ax3.grid(True, alpha=0.3, axis='y')
            ax3.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax3.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 4: Strip Ratio
        ax4 = axes[1, 1]
        plotted = False
        if 'Overburden_RGM' in self.clean_df.columns and 'Ore_Mined_RGM' in self.clean_df.columns:
            strip_ratio_rgm = self.clean_df['Overburden_RGM'] / (self.clean_df['Ore_Mined_RGM'] + 0.001)
            strip_ratio_rgm = strip_ratio_rgm.replace([np.inf, -np.inf], np.nan)
            ax4.plot(self.clean_df.index, strip_ratio_rgm, marker='o', 
                    label='RGM Strip Ratio', linewidth=2, markersize=4)
            plotted = True
            
        if 'Overburden_Sar' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            strip_ratio_sar = self.clean_df['Overburden_Sar'] / (self.clean_df['Ore_Mined_Sar'] + 0.001)
            strip_ratio_sar = strip_ratio_sar.replace([np.inf, -np.inf], np.nan)
            ax4.plot(self.clean_df.index, strip_ratio_sar, marker='s', 
                    label='Sar Strip Ratio', linewidth=2, markersize=4)
            plotted = True
            
        if plotted:
            ax4.set_title('Stripping Ratio Trends', fontweight='bold')
            ax4.set_xlabel('Date')
            ax4.set_ylabel('Strip Ratio')
            ax4.legend()
            ax4.grid(True, alpha=0.3)
            ax4.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax4.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45)
        
        plt.tight_layout()
        filename = os.path.join(self.temp_dir, 'production_overview.png')
        plt.savefig(filename, dpi=100, bbox_inches='tight')
        plt.close()
        return filename
    
    def _create_efficiency_charts(self):
        """Create efficiency analysis charts"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Efficiency Analysis', fontsize=16, fontweight='bold')
        
        # Chart 1: Fleet vs Production
        ax1 = axes[0, 0]
        if 'Active_Fleet_Count_Aprox' in self.clean_df.columns:
            total_production = pd.Series(0, index=self.clean_df.index)
            for col in ['Ore_Mined_RGM', 'Ore_Mined_Sar']:
                if col in self.clean_df.columns:
                    total_production += self.clean_df[col]
            
            ax1_twin = ax1.twinx()
            ax1.bar(self.clean_df.index, total_production, alpha=0.5, 
                   color='skyblue', label='Total Ore Production')
            ax1_twin.plot(self.clean_df.index, self.clean_df['Active_Fleet_Count_Aprox'], 
                         color='red', marker='o', linewidth=2, markersize=4, 
                         label='Fleet Count')
            
            ax1.set_xlabel('Date')
            ax1.set_ylabel('Ore Production (kt)', color='blue')
            ax1_twin.set_ylabel('Fleet Count', color='red')
            ax1.set_title('Fleet Utilization vs Production', fontweight='bold')
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
            ax1.grid(True, alpha=0.3)
        
        # Chart 2: Diesel Consumption
        ax2 = axes[0, 1]
        if 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
            diesel_ml = self.clean_df['Liter_of_Diesel_Consumed'] / 1000000
            ax2.plot(self.clean_df.index, diesel_ml, 
                    marker='o', color='green', linewidth=2, markersize=4)
            ax2.fill_between(self.clean_df.index, diesel_ml, alpha=0.3, color='green')
            ax2.set_title('Diesel Consumption Trend', fontweight='bold')
            ax2.set_xlabel('Date')
            ax2.set_ylabel('Diesel (Million Liters)')
            ax2.grid(True, alpha=0.3)
            ax2.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 3: Productivity per Fleet Unit
        ax3 = axes[1, 0]
        if 'Active_Fleet_Count_Aprox' in self.clean_df.columns:
            total_material = pd.Series(0, index=self.clean_df.index)
            for col in self.clean_df.columns:
                if 'Ore_Mined' in col or 'Overburden' in col:
                    total_material += self.clean_df[col]
            
            productivity = total_material / (self.clean_df['Active_Fleet_Count_Aprox'] + 0.001)
            ax3.plot(self.clean_df.index, productivity, marker='o', 
                    color='purple', linewidth=2, markersize=4)
            ax3.set_title('Productivity per Fleet Unit', fontweight='bold')
            ax3.set_xlabel('Date')
            ax3.set_ylabel('Material per Unit (kt)')
            ax3.grid(True, alpha=0.3)
            ax3.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax3.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 4: Fuel Efficiency
        ax4 = axes[1, 1]
        if 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
            total_material = pd.Series(0, index=self.clean_df.index)
            for col in self.clean_df.columns:
                if 'Ore_Mined' in col or 'Overburden' in col:
                    total_material += self.clean_df[col]
            
            fuel_efficiency = self.clean_df['Liter_of_Diesel_Consumed'] / (total_material + 0.001)
            fuel_efficiency = fuel_efficiency.replace([np.inf, -np.inf], np.nan)
            ax4.plot(self.clean_df.index, fuel_efficiency, marker='o', 
                    color='orange', linewidth=2, markersize=4)
            ax4.set_title('Fuel Efficiency (L/kt)', fontweight='bold')
            ax4.set_xlabel('Date')
            ax4.set_ylabel('Liters per kt')
            ax4.grid(True, alpha=0.3)
            ax4.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax4.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45)
        
        plt.tight_layout()
        filename = os.path.join(self.temp_dir, 'efficiency_analysis.png')
        plt.savefig(filename, dpi=100, bbox_inches='tight')
        plt.close()
        return filename
    
    def _create_comparative_charts(self):
        """Create comparative analysis charts"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Comparative Analysis: RGM vs Sar', fontsize=16, fontweight='bold')
        
        # Chart 1: Production Share Pie Chart
        ax1 = axes[0, 0]
        if 'Ore_Mined_RGM' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            rgm_total = self.clean_df['Ore_Mined_RGM'].sum()
            sar_total = self.clean_df['Ore_Mined_Sar'].sum()
            
            if rgm_total > 0 or sar_total > 0:
                ax1.pie([rgm_total, sar_total], labels=['RGM', 'Sar'], 
                       autopct='%1.1f%%', startangle=90, 
                       colors=['#3498db', '#e74c3c'])
                ax1.set_title('Total Ore Production Share', fontweight='bold')
        
        # Chart 2: Monthly Comparison Bar Chart
        ax2 = axes[0, 1]
        if 'Ore_Mined_RGM' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            # Sample every 3rd month for clarity
            sample_data = self.clean_df.iloc[::3]
            x = np.arange(len(sample_data))
            width = 0.35
            
            ax2.bar(x - width/2, sample_data['Ore_Mined_RGM'], width, 
                   label='RGM', color='#3498db', alpha=0.8)
            ax2.bar(x + width/2, sample_data['Ore_Mined_Sar'], width, 
                   label='Sar', color='#e74c3c', alpha=0.8)
            
            ax2.set_xlabel('Date')
            ax2.set_ylabel('Ore Mined (kt)')
            ax2.set_title('Ore Production Comparison', fontweight='bold')
            ax2.set_xticks(x)
            ax2.set_xticklabels([d.strftime('%b-%y') for d in sample_data.index], rotation=45)
            ax2.legend()
            ax2.grid(True, alpha=0.3, axis='y')
        
        # Chart 3: Overburden Share Pie Chart
        ax3 = axes[1, 0]
        if 'Overburden_RGM' in self.clean_df.columns and 'Overburden_Sar' in self.clean_df.columns:
            rgm_ob = self.clean_df['Overburden_RGM'].sum()
            sar_ob = self.clean_df['Overburden_Sar'].sum()
            
            if rgm_ob > 0 or sar_ob > 0:
                ax3.pie([rgm_ob, sar_ob], labels=['RGM', 'Sar'], 
                       autopct='%1.1f%%', startangle=90, 
                       colors=['#9b59b6', '#f39c12'])
                ax3.set_title('Total Overburden Share', fontweight='bold')
        
        # Chart 4: Performance Metrics Comparison
        ax4 = axes[1, 1]
        metrics = []
        rgm_values = []
        sar_values = []
        
        if 'Ore_Mined_RGM' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            # Average production
            metrics.append('Avg Ore\n(kt/month)')
            rgm_values.append(self.clean_df['Ore_Mined_RGM'].mean())
            sar_values.append(self.clean_df['Ore_Mined_Sar'].mean())
            
            # Max production
            metrics.append('Max Ore\n(kt)')
            rgm_values.append(self.clean_df['Ore_Mined_RGM'].max())
            sar_values.append(self.clean_df['Ore_Mined_Sar'].max())
            
            # Total production
            metrics.append('Total Ore\n(kt/1000)')
            rgm_values.append(self.clean_df['Ore_Mined_RGM'].sum()/1000)
            sar_values.append(self.clean_df['Ore_Mined_Sar'].sum()/1000)
        
        if metrics:
            x = np.arange(len(metrics))
            width = 0.35
            
            ax4.bar(x - width/2, rgm_values, width, label='RGM', color='#3498db')
            ax4.bar(x + width/2, sar_values, width, label='Sar', color='#e74c3c')
            
            ax4.set_ylabel('Values')
            ax4.set_title('Performance Metrics', fontweight='bold')
            ax4.set_xticks(x)
            ax4.set_xticklabels(metrics)
            ax4.legend()
            ax4.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        filename = os.path.join(self.temp_dir, 'comparative_analysis.png')
        plt.savefig(filename, dpi=100, bbox_inches='tight')
        plt.close()
        return filename
    
    def _create_trend_charts(self):
        """Create trend analysis charts"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Trend Analysis', fontsize=16, fontweight='bold')
        
        # Chart 1: Moving Averages
        ax1 = axes[0, 0]
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            ma_3 = self.clean_df['Ore_Mined_RGM'].rolling(window=3, center=True).mean()
            ma_6 = self.clean_df['Ore_Mined_RGM'].rolling(window=6, center=True).mean()
            
            ax1.plot(self.clean_df.index, self.clean_df['Ore_Mined_RGM'], 
                    alpha=0.3, label='Actual', linewidth=1, color='gray')
            ax1.plot(self.clean_df.index, ma_3, label='3-Month MA', 
                    linewidth=2, color='red')
            ax1.plot(self.clean_df.index, ma_6, label='6-Month MA', 
                    linewidth=2, color='green')
            
            ax1.set_title('RGM Production - Moving Averages', fontweight='bold')
            ax1.set_xlabel('Date')
            ax1.set_ylabel('Ore Mined (kt)')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=6))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 2: Year-over-Year Comparison
        ax2 = axes[0, 1]
        yearly_data = {}
        
        for col in ['Ore_Mined_RGM', 'Ore_Mined_Sar']:
            if col in self.clean_df.columns:
                yearly = self.clean_df[col].groupby(self.clean_df.index.year).sum()
                for year, value in yearly.items():
                    if year not in yearly_data:
                        yearly_data[year] = {}
                    yearly_data[year][col] = value
        
        if yearly_data:
            years = sorted(yearly_data.keys())
            rgm_yearly = [yearly_data[y].get('Ore_Mined_RGM', 0) for y in years]
            sar_yearly = [yearly_data[y].get('Ore_Mined_Sar', 0) for y in years]
            
            x = np.arange(len(years))
            width = 0.35
            
            ax2.bar(x - width/2, rgm_yearly, width, label='RGM', color='#3498db')
            ax2.bar(x + width/2, sar_yearly, width, label='Sar', color='#e74c3c')
            
            ax2.set_xlabel('Year')
            ax2.set_ylabel('Total Ore (kt)')
            ax2.set_title('Yearly Production Comparison', fontweight='bold')
            ax2.set_xticks(x)
            ax2.set_xticklabels(years)
            ax2.legend()
            ax2.grid(True, alpha=0.3, axis='y')
        
        # Chart 3: Fleet vs Diesel Correlation
        ax3 = axes[1, 0]
        if 'Active_Fleet_Count_Aprox' in self.clean_df.columns and 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
            fleet = self.clean_df['Active_Fleet_Count_Aprox']
            diesel = self.clean_df['Liter_of_Diesel_Consumed'] / 1000000
            
            # Remove zeros and NaNs for correlation
            mask = (fleet > 0) & (diesel > 0)
            fleet_clean = fleet[mask]
            diesel_clean = diesel[mask]
            
            if len(fleet_clean) > 0:
                scatter = ax3.scatter(fleet_clean, diesel_clean, 
                                    c=range(len(fleet_clean)), cmap='viridis', 
                                    alpha=0.6, s=50)
                
                # Add trend line
                z = np.polyfit(fleet_clean, diesel_clean, 1)
                p = np.poly1d(z)
                ax3.plot(sorted(fleet_clean), p(sorted(fleet_clean)), 
                        "r--", alpha=0.8, label=f'Trend')
                
                ax3.set_xlabel('Fleet Count')
                ax3.set_ylabel('Diesel (Million L)')
                ax3.set_title('Fleet Count vs Diesel Consumption', fontweight='bold')
                ax3.legend()
                ax3.grid(True, alpha=0.3)
                plt.colorbar(scatter, ax=ax3, label='Time')
        
        # Chart 4: Monthly Seasonality
        ax4 = axes[1, 1]
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            monthly_avg = self.clean_df['Ore_Mined_RGM'].groupby(self.clean_df.index.month).mean()
            
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                     'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            
            bars = ax4.bar(range(1, 13), monthly_avg.reindex(range(1, 13), fill_value=0), 
                          color='steelblue', alpha=0.7)
            
            # Add average line
            avg_line = monthly_avg.mean()
            ax4.axhline(y=avg_line, color='red', linestyle='--', 
                       label=f'Overall Avg: {avg_line:.1f}')
            
            ax4.set_xlabel('Month')
            ax4.set_ylabel('Average Ore (kt)')
            ax4.set_title('RGM Production - Monthly Seasonality', fontweight='bold')
            ax4.set_xticks(range(1, 13))
            ax4.set_xticklabels(months, rotation=45)
            ax4.legend()
            ax4.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        filename = os.path.join(self.temp_dir, 'trend_analysis.png')
        plt.savefig(filename, dpi=100, bbox_inches='tight')
        plt.close()
        return filename
    
    def create_excel_with_charts(self, output_file=None):
        """Create Excel file with embedded charts"""
        if output_file is None:
            base_name = os.path.splitext(self.input_file)[0]
            output_file = f"{base_name}_with_charts.xlsx"
        
        print("\n📝 Creating Excel file with charts...")
        
        # Generate charts first
        chart_files = self.generate_charts()
        
        # Create new workbook
        wb = Workbook()
        
        # Add data sheet
        ws_data = wb.active
        ws_data.title = "Processed_Data"
        
        # Write processed data
        for r in dataframe_to_rows(self.clean_df.reset_index(), index=False, header=True):
            ws_data.append(r)
        
        # Format headers
        for cell in ws_data[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="366092")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust columns
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # Add chart sheets
        for sheet_name, chart_file in chart_files:
            ws_chart = wb.create_sheet(title=sheet_name)
            
            # Add title
            ws_chart['A1'] = sheet_name
            ws_chart['A1'].font = Font(size=16, bold=True)
            
            # Insert image
            img = Image(chart_file)
            img.width = 1100  # Adjust size
            img.height = 750
            ws_chart.add_image(img, 'A3')
        
        # Add summary sheet
        ws_summary = wb.create_sheet(title="Summary", index=0)
        
        # Create summary statistics
        summary_data = [
            ['Mining Data Analysis Summary'],
            [''],
            ['Metric', 'RGM', 'Sar'],
            [''],
        ]
        
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            summary_data.append(['Total Ore (kt)', 
                               f"{self.clean_df['Ore_Mined_RGM'].sum():,.0f}",
                               f"{self.clean_df.get('Ore_Mined_Sar', pd.Series([0])).sum():,.0f}"])
            summary_data.append(['Average Ore/Month (kt)', 
                               f"{self.clean_df['Ore_Mined_RGM'].mean():,.1f}",
                               f"{self.clean_df.get('Ore_Mined_Sar', pd.Series([0])).mean():,.1f}"])
            summary_data.append(['Max Ore (kt)', 
                               f"{self.clean_df['Ore_Mined_RGM'].max():,.1f}",
                               f"{self.clean_df.get('Ore_Mined_Sar', pd.Series([0])).max():,.1f}"])
        
        if 'Overburden_RGM' in self.clean_df.columns:
            summary_data.append(['Total Overburden (kt)', 
                               f"{self.clean_df['Overburden_RGM'].sum():,.0f}",
                               f"{self.clean_df.get('Overburden_Sar', pd.Series([0])).sum():,.0f}"])
        
        summary_data.append([''])
        summary_data.append(['Date Range', f"{self.clean_df.index[0].strftime('%b %Y')} to {self.clean_df.index[-1].strftime('%b %Y')}"])
        summary_data.append(['Total Periods', f"{len(self.clean_df)} months"])
        
        if 'Active_Fleet_Count_Aprox' in self.clean_df.columns:
            summary_data.append(['Avg Fleet Count', f"{self.clean_df['Active_Fleet_Count_Aprox'].mean():,.0f}"])
        
        if 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
            summary_data.append(['Total Diesel (Million L)', 
                               f"{self.clean_df['Liter_of_Diesel_Consumed'].sum()/1000000:,.1f}"])
        
        # Write summary data
        for row in summary_data:
            ws_summary.append(row)
        
        # Format summary sheet
        ws_summary['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws_summary['A1'].fill = PatternFill("solid", fgColor="366092")
        ws_summary.merge_cells('A1:C1')
        
        ws_summary['A3'].font = Font(bold=True)
        ws_summary['B3'].font = Font(bold=True)
        ws_summary['C3'].font = Font(bold=True)
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 15
        
        # Save workbook
        wb.save(output_file)
        
        print(f"✅ Excel file saved: {output_file}")
        
        # Clean up temp files
        for _, chart_file in chart_files:
            try:
                os.remove(chart_file)
            except:
                pass
        
        return output_file


def main():
    """Main function"""
    print("=" * 60)
    print("  Mining Data Chart Generator")
    print("  Creates Excel file with embedded charts")
    print("=" * 60)
    
    # Get input file
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = input("\nEnter Excel file path: ").strip()
        if not input_file:
            print("❌ No file specified. Using sample data...")
            input_file = "sample_mining_data.xlsx"
    
    # Check file exists
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        return
    
    # Get sheet name
    sheet_name = None
    if len(sys.argv) > 2:
        sheet_name = sys.argv[2]
    
    try:
        # Create generator
        generator = MiningDataChartGenerator(input_file, sheet_name)
        
        # Load and process data
        generator.load_data()
        generator.process_data()
        
        # Create Excel with charts
        output_file = generator.create_excel_with_charts()
        
        print("\n" + "=" * 60)
        print("  ✅ SUCCESS! Charts created and embedded in Excel")
        print(f"  📁 Output: {output_file}")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
