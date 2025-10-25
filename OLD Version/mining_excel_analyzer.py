import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import (
    LineChart, BarChart, PieChart, ScatterChart,
    Reference, Series, DataPoint
)
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

class MiningExcelAnalyzer:
    def __init__(self, input_file, sheet_name=None):
        """Initialize the analyzer with an Excel file"""
        self.input_file = input_file
        self.sheet_name = sheet_name
        self.df = None
        self.clean_df = None
        self.wb = None
        
    def load_and_process_data(self):
        """Load and process the Excel data"""
        print(f"Loading file: {self.input_file}")
        
        # Read the Excel file
        if self.sheet_name:
            self.df = pd.read_excel(self.input_file, sheet_name=self.sheet_name)
        else:
            self.df = pd.read_excel(self.input_file)
            
        print(f"Data loaded: {self.df.shape[0]} rows, {self.df.shape[1]} columns")
        
        # Process the data
        self._process_data()
        
    def _process_data(self):
        """Process and clean the data"""
        try:
            # Find the header row
            header_row = None
            for idx, row in self.df.iterrows():
                if any(isinstance(val, str) and 'ene-' in str(val).lower() for val in row.values):
                    header_row = idx
                    break
            
            # Set the header row
            if header_row is not None:
                self.df.columns = self.df.iloc[header_row]
                self.df = self.df.iloc[header_row + 1:].reset_index(drop=True)
            
            # Clean column names
            self.df.columns = [str(col).strip() if pd.notna(col) else f'Col_{i}' 
                              for i, col in enumerate(self.df.columns)]
            
            # Create a cleaner dataframe
            data_dict = {}
            
            # Extract metrics
            for idx, row in self.df.iterrows():
                if pd.notna(row.iloc[0]):
                    metric_name = str(row.iloc[0]).strip()
                    if row.iloc[1:3].notna().any():  # Has subcategory
                        subcategory = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
                        key = f"{metric_name}_{subcategory}".replace(' ', '_')
                    else:
                        key = metric_name.replace(' ', '_')
                    
                    # Get the data values
                    values = []
                    dates = []
                    for col in self.df.columns[3:]:
                        if '-' in str(col):  # This is a date column
                            dates.append(str(col))
                            val = row[col]
                            if pd.notna(val):
                                if isinstance(val, str):
                                    val = val.replace('.', '').replace(',', '.')
                                try:
                                    values.append(float(val))
                                except:
                                    values.append(0)
                            else:
                                values.append(0)
                    
                    if values and not all(v == 0 for v in values):
                        data_dict[key] = values
            
            # Get date columns
            date_cols = [col for col in self.df.columns[3:] if '-' in str(col)]
            
            # Create clean dataframe
            self.clean_df = pd.DataFrame(data_dict)
            
            # Parse dates safely
            parsed_dates = []
            months_map = {'ene': '01', 'feb': '02', 'mar': '03', 'abr': '04', 
                         'may': '05', 'jun': '06', 'jul': '07', 'ago': '08',
                         'sep': '09', 'oct': '10', 'nov': '11', 'dic': '12'}
            
            for date_str in date_cols[:len(self.clean_df)]:
                try:
                    parts = str(date_str).split('-')
                    if len(parts) == 2:
                        month = months_map.get(parts[0].lower(), '01')
                        year = '20' + parts[1] if len(parts[1]) == 2 else parts[1]
                        parsed_dates.append(f"{year}-{month}-01")
                    else:
                        parsed_dates.append(None)
                except:
                    parsed_dates.append(None)
            
            # Add parsed dates
            self.clean_df['Date'] = parsed_dates
            
            # Convert to datetime, handling NaT values
            self.clean_df['Date'] = pd.to_datetime(self.clean_df['Date'], errors='coerce')
            
            # Remove rows with invalid dates
            self.clean_df = self.clean_df[self.clean_df['Date'].notna()]
            
            # Set date as index and sort
            self.clean_df = self.clean_df.set_index('Date').sort_index()
            
            # Fill NaN values with 0 for calculations
            self.clean_df = self.clean_df.fillna(0)
            
            print(f"Processed data: {len(self.clean_df)} time periods, {len(self.clean_df.columns)} metrics")
            print(f"Metrics found: {', '.join(self.clean_df.columns[:10])}")
            
        except Exception as e:
            print(f"Error processing data: {str(e)}")
            raise
            
    def create_excel_with_charts(self, output_file=None):
        """Create a new Excel file with data and charts"""
        if output_file is None:
            base_name = os.path.splitext(self.input_file)[0]
            output_file = f"{base_name}_with_charts.xlsx"
            
        print(f"\nCreating Excel file with charts: {output_file}")
        
        # Create a new Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # 1. Write original data
            original_df = pd.read_excel(self.input_file, sheet_name=self.sheet_name) if self.sheet_name else pd.read_excel(self.input_file)
            original_df.to_excel(writer, sheet_name='Original_Data', index=False)
            
            # 2. Write processed data
            self.clean_df.to_excel(writer, sheet_name='Processed_Data')
            
            # 3. Create summary statistics
            self._create_summary_sheet(writer)
            
            # 4. Create production analysis
            self._create_production_analysis(writer)
            
            # 5. Create efficiency analysis
            self._create_efficiency_analysis(writer)
            
            # 6. Create comparative analysis
            self._create_comparative_analysis(writer)
            
            # 7. Create trend analysis
            self._create_trend_analysis(writer)
            
        print(f"✅ Excel file created successfully: {output_file}")
        return output_file
    
    def _create_summary_sheet(self, writer):
        """Create a summary statistics sheet"""
        summary_data = []
        
        for col in self.clean_df.columns:
            if self.clean_df[col].dtype in ['float64', 'int64']:
                summary_data.append({
                    'Metric': col,
                    'Mean': self.clean_df[col].mean(),
                    'Std Dev': self.clean_df[col].std(),
                    'Min': self.clean_df[col].min(),
                    'Max': self.clean_df[col].max(),
                    'Total': self.clean_df[col].sum()
                })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary_Statistics', index=False)
        
        # Format the sheet
        worksheet = writer.sheets['Summary_Statistics']
        self._format_header(worksheet)
        
    def _create_production_analysis(self, writer):
        """Create production analysis sheet with data and charts"""
        production_data = pd.DataFrame()
        production_data['Date'] = self.clean_df.index
        
        # Add production metrics
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            production_data['Ore_RGM'] = self.clean_df['Ore_Mined_RGM'].values
        if 'Ore_Mined_Sar' in self.clean_df.columns:
            production_data['Ore_Sar'] = self.clean_df['Ore_Mined_Sar'].values
        if 'Overburden_RGM' in self.clean_df.columns:
            production_data['Overburden_RGM'] = self.clean_df['Overburden_RGM'].values
        if 'Overburden_Sar' in self.clean_df.columns:
            production_data['Overburden_Sar'] = self.clean_df['Overburden_Sar'].values
            
        # Calculate strip ratios
        if 'Ore_RGM' in production_data.columns and 'Overburden_RGM' in production_data.columns:
            production_data['Strip_Ratio_RGM'] = production_data['Overburden_RGM'] / (production_data['Ore_RGM'] + 0.001)
        if 'Ore_Sar' in production_data.columns and 'Overburden_Sar' in production_data.columns:
            production_data['Strip_Ratio_Sar'] = production_data['Overburden_Sar'] / (production_data['Ore_Sar'] + 0.001)
        
        # Write to Excel
        production_data.to_excel(writer, sheet_name='Production_Analysis', index=False)
        
        # Format the sheet
        worksheet = writer.sheets['Production_Analysis']
        self._format_header(worksheet)
        
        # Add charts using openpyxl
        wb = writer.book
        ws = wb['Production_Analysis']
        
        # Chart 1: Ore Production Comparison
        if 'Ore_RGM' in production_data.columns and 'Ore_Sar' in production_data.columns:
            chart1 = LineChart()
            chart1.title = "Ore Production Comparison"
            chart1.x_axis.title = "Date"
            chart1.y_axis.title = "Ore Mined (kt)"
            
            # Add data
            data1 = Reference(ws, min_col=2, min_row=1, max_row=len(production_data)+1)
            data2 = Reference(ws, min_col=3, min_row=1, max_row=len(production_data)+1)
            dates = Reference(ws, min_col=1, min_row=2, max_row=len(production_data)+1)
            
            chart1.add_data(data1, titles_from_data=True)
            chart1.add_data(data2, titles_from_data=True)
            chart1.set_categories(dates)
            
            ws.add_chart(chart1, "I2")
        
        # Chart 2: Strip Ratio Trends
        if 'Strip_Ratio_RGM' in production_data.columns:
            chart2 = LineChart()
            chart2.title = "Strip Ratio Trends"
            chart2.x_axis.title = "Date"
            chart2.y_axis.title = "Strip Ratio"
            
            col_idx = production_data.columns.get_loc('Strip_Ratio_RGM') + 1
            data = Reference(ws, min_col=col_idx, min_row=1, max_row=len(production_data)+1)
            dates = Reference(ws, min_col=1, min_row=2, max_row=len(production_data)+1)
            
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(dates)
            
            ws.add_chart(chart2, "I20")
            
    def _create_efficiency_analysis(self, writer):
        """Create efficiency analysis sheet"""
        efficiency_data = pd.DataFrame()
        efficiency_data['Date'] = self.clean_df.index
        
        # Add efficiency metrics
        if 'Active_Fleet_Count_(Aprox)' in self.clean_df.columns:
            efficiency_data['Fleet_Count'] = self.clean_df['Active_Fleet_Count_(Aprox)'].values
            
        if 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
            efficiency_data['Diesel_ML'] = self.clean_df['Liter_of_Diesel_Consumed'].values / 1000000
            
        # Calculate total material
        total_material = pd.Series(0, index=self.clean_df.index)
        for col in ['Ore_Mined_RGM', 'Ore_Mined_Sar', 'Overburden_RGM', 'Overburden_Sar']:
            if col in self.clean_df.columns:
                total_material += self.clean_df[col]
        efficiency_data['Total_Material_kt'] = total_material.values
        
        # Calculate productivity metrics
        if 'Fleet_Count' in efficiency_data.columns and 'Total_Material_kt' in efficiency_data.columns:
            efficiency_data['Productivity_per_Unit'] = efficiency_data['Total_Material_kt'] / (efficiency_data['Fleet_Count'] + 0.001)
            
        if 'Diesel_ML' in efficiency_data.columns and 'Total_Material_kt' in efficiency_data.columns:
            efficiency_data['Fuel_Efficiency_L_per_kt'] = (efficiency_data['Diesel_ML'] * 1000000) / (efficiency_data['Total_Material_kt'] + 0.001)
        
        # Write to Excel
        efficiency_data.to_excel(writer, sheet_name='Efficiency_Analysis', index=False)
        
        # Format the sheet
        worksheet = writer.sheets['Efficiency_Analysis']
        self._format_header(worksheet)
        
    def _create_comparative_analysis(self, writer):
        """Create comparative analysis sheet"""
        comparison_data = {}
        
        # Calculate totals and averages for RGM vs Sar
        metrics = ['Total Ore (kt)', 'Avg Ore/Month (kt)', 'Total Overburden (kt)', 
                  'Avg Overburden/Month (kt)', 'Avg Strip Ratio']
        
        rgm_values = []
        sar_values = []
        
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            rgm_values.append(self.clean_df['Ore_Mined_RGM'].sum())
            rgm_values.append(self.clean_df['Ore_Mined_RGM'].mean())
        else:
            rgm_values.extend([0, 0])
            
        if 'Ore_Mined_Sar' in self.clean_df.columns:
            sar_values.append(self.clean_df['Ore_Mined_Sar'].sum())
            sar_values.append(self.clean_df['Ore_Mined_Sar'].mean())
        else:
            sar_values.extend([0, 0])
            
        if 'Overburden_RGM' in self.clean_df.columns:
            rgm_values.append(self.clean_df['Overburden_RGM'].sum())
            rgm_values.append(self.clean_df['Overburden_RGM'].mean())
        else:
            rgm_values.extend([0, 0])
            
        if 'Overburden_Sar' in self.clean_df.columns:
            sar_values.append(self.clean_df['Overburden_Sar'].sum())
            sar_values.append(self.clean_df['Overburden_Sar'].mean())
        else:
            sar_values.extend([0, 0])
            
        # Calculate strip ratios
        if 'Overburden_RGM' in self.clean_df.columns and 'Ore_Mined_RGM' in self.clean_df.columns:
            strip_rgm = (self.clean_df['Overburden_RGM'] / (self.clean_df['Ore_Mined_RGM'] + 0.001)).mean()
            rgm_values.append(strip_rgm)
        else:
            rgm_values.append(0)
            
        if 'Overburden_Sar' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            strip_sar = (self.clean_df['Overburden_Sar'] / (self.clean_df['Ore_Mined_Sar'] + 0.001)).mean()
            sar_values.append(strip_sar)
        else:
            sar_values.append(0)
        
        comparison_df = pd.DataFrame({
            'Metric': metrics,
            'RGM': rgm_values,
            'Sar': sar_values
        })
        
        comparison_df.to_excel(writer, sheet_name='Comparative_Analysis', index=False)
        
        # Format the sheet
        worksheet = writer.sheets['Comparative_Analysis']
        self._format_header(worksheet)
        
    def _create_trend_analysis(self, writer):
        """Create trend analysis sheet"""
        trend_data = pd.DataFrame()
        trend_data['Date'] = self.clean_df.index
        
        # Add key metrics for trend analysis
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            trend_data['Ore_RGM'] = self.clean_df['Ore_Mined_RGM'].values
            trend_data['Ore_RGM_MA3'] = self.clean_df['Ore_Mined_RGM'].rolling(window=3, center=True).mean().values
            trend_data['Ore_RGM_MA6'] = self.clean_df['Ore_Mined_RGM'].rolling(window=6, center=True).mean().values
        
        # Yearly aggregation
        yearly_summary = []
        for year in self.clean_df.index.year.unique():
            year_data = self.clean_df[self.clean_df.index.year == year]
            
            summary = {'Year': year}
            if 'Ore_Mined_RGM' in self.clean_df.columns:
                summary['Total_Ore_RGM'] = year_data['Ore_Mined_RGM'].sum()
            if 'Ore_Mined_Sar' in self.clean_df.columns:
                summary['Total_Ore_Sar'] = year_data['Ore_Mined_Sar'].sum()
            if 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
                summary['Total_Diesel_ML'] = year_data['Liter_of_Diesel_Consumed'].sum() / 1000000
            
            yearly_summary.append(summary)
        
        # Write trend data
        trend_data.to_excel(writer, sheet_name='Trend_Analysis', index=False)
        
        # Write yearly summary
        if yearly_summary:
            yearly_df = pd.DataFrame(yearly_summary)
            yearly_df.to_excel(writer, sheet_name='Yearly_Summary', index=False)
            
            # Format the sheet
            worksheet = writer.sheets['Yearly_Summary']
            self._format_header(worksheet)
        
        # Format the trend sheet
        worksheet = writer.sheets['Trend_Analysis']
        self._format_header(worksheet)
        
    def _format_header(self, worksheet):
        """Format header row in worksheet"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """Main function to run the analyzer"""
    print("=" * 60)
    print("Mining Data Excel Analyzer")
    print("=" * 60)
    
    # Get input file
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = input("Enter the path to your Excel file: ").strip()
    
    # Check if file exists
    if not os.path.exists(input_file):
        print(f"❌ Error: File '{input_file}' not found!")
        return
    
    # Get sheet name (optional)
    sheet_name = None
    if len(sys.argv) > 2:
        sheet_name = sys.argv[2]
    else:
        response = input("Enter sheet name (press Enter to use first sheet): ").strip()
        if response:
            sheet_name = response
    
    try:
        # Create analyzer
        analyzer = MiningExcelAnalyzer(input_file, sheet_name)
        
        # Load and process data
        print("\n📊 Processing data...")
        analyzer.load_and_process_data()
        
        # Create Excel with charts
        print("\n📈 Creating charts and analysis...")
        output_file = analyzer.create_excel_with_charts()
        
        print("\n" + "=" * 60)
        print("✅ Analysis complete!")
        print(f"📁 Output file: {output_file}")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
