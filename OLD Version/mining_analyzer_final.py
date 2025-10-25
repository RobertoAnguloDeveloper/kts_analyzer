#!/usr/bin/env python3
"""
Mining Data Analyzer - Complete All-in-One Solution
Analyzes mining data from Excel and creates embedded charts
Works in both GUI and Command-Line modes
Version: FINAL
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys
import tempfile
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Try to import tkinter for GUI mode
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    import threading
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False
    print("Note: GUI mode not available. Running in command-line mode.")

# Set matplotlib to non-interactive backend
plt.switch_backend('Agg')

# Set style
try:
    plt.style.use('seaborn-v0_8-darkgrid')
except:
    try:
        plt.style.use('seaborn-darkgrid')
    except:
        pass  # Use default style


class MiningDataProcessor:
    """Core data processor for mining data analysis"""
    
    def __init__(self, input_file, sheet_name=None, output_file=None):
        self.input_file = input_file
        self.sheet_name = sheet_name
        self.output_file = output_file
        self.df = None
        self.clean_df = None
        self.temp_dir = tempfile.mkdtemp()
        self.chart_files = []
        
    def process(self):
        """Main processing pipeline"""
        print("📂 Loading data...")
        self.load_data()
        
        print("🔧 Processing data...")
        self.process_data()
        
        print("📊 Generating charts...")
        self.create_charts()
        
        print("📝 Creating Excel file...")
        output = self.create_excel_with_charts()
        
        print(f"✅ Success! Output file: {output}")
        return output
        
    def load_data(self):
        """Load Excel data"""
        if self.sheet_name:
            self.df = pd.read_excel(self.input_file, sheet_name=self.sheet_name)
        else:
            excel_file = pd.ExcelFile(self.input_file)
            sheets = excel_file.sheet_names
            print(f"   Found sheets: {', '.join(sheets)}")
            self.df = pd.read_excel(self.input_file, sheet_name=sheets[0])
            print(f"   Using sheet: {sheets[0]}")
            
    def process_data(self):
        """Process and clean the data"""
        # Find header row with date columns
        header_row = None
        for idx in range(min(10, len(self.df))):
            row_values = self.df.iloc[idx].astype(str).values
            if any(any(month in val.lower() for month in ['ene-', 'feb-', 'mar-', 'abr-', 'may-', 'jun-', 
                                                          'jul-', 'ago-', 'sep-', 'oct-', 'nov-', 'dic-'])
                  for val in row_values):
                header_row = idx
                break
        
        # Set proper column names
        if header_row is not None and header_row > 0:
            self.df.columns = self.df.iloc[header_row]
            self.df = self.df.iloc[header_row + 1:].reset_index(drop=True)
        
        # Clean column names
        self.df.columns = [str(col).strip() if pd.notna(col) else f'Col_{i}' 
                          for i, col in enumerate(self.df.columns)]
        
        # Get date columns
        date_columns = [col for col in self.df.columns[3:] if '-' in str(col)]
        print(f"   Found {len(date_columns)} date columns")
        
        # Extract data into dictionary
        data_dict = {}
        
        for idx, row in self.df.iterrows():
            if pd.notna(row.iloc[0]):
                metric_name = str(row.iloc[0]).strip()
                
                # Handle subcategory
                if len(row) > 1 and pd.notna(row.iloc[1]):
                    subcategory = str(row.iloc[1]).strip()
                    if subcategory.lower() not in ['nan', '', 'none']:
                        key = f"{metric_name}_{subcategory}"
                    else:
                        key = metric_name
                else:
                    key = metric_name
                
                # Clean key
                key = key.replace(' ', '_').replace('(', '').replace(')', '')
                
                # Extract values
                values = []
                for date_col in date_columns:
                    if date_col in self.df.columns:
                        val = row[date_col]
                        if pd.notna(val):
                            if isinstance(val, str):
                                # Clean number format
                                val = val.replace(',', '').replace(' ', '')
                                if '.' in val and val.count('.') > 1:
                                    val = val.replace('.', '')
                            try:
                                values.append(float(val))
                            except:
                                values.append(0)
                        else:
                            values.append(0)
                
                if values and not all(v == 0 for v in values):
                    data_dict[key] = values
        
        # Create clean dataframe
        self.clean_df = pd.DataFrame(data_dict)
        
        # Parse dates
        parsed_dates = self._parse_dates(date_columns[:len(self.clean_df)])
        self.clean_df['Date'] = parsed_dates
        
        # Convert to datetime
        self.clean_df['Date'] = pd.to_datetime(self.clean_df['Date'], errors='coerce')
        self.clean_df = self.clean_df.dropna(subset=['Date'])
        self.clean_df = self.clean_df.set_index('Date').sort_index()
        
        print(f"   Processed {len(self.clean_df)} periods with {len(self.clean_df.columns)} metrics")
        if len(self.clean_df) > 0:
            print(f"   Date range: {self.clean_df.index[0].strftime('%b %Y')} to {self.clean_df.index[-1].strftime('%b %Y')}")
        
    def _parse_dates(self, date_strings):
        """Parse Spanish/English month abbreviations"""
        months = {
            'ene': '01', 'jan': '01', 'enero': '01',
            'feb': '02', 'febrero': '02',
            'mar': '03', 'marzo': '03',
            'abr': '04', 'apr': '04', 'abril': '04',
            'may': '05', 'mayo': '05',
            'jun': '06', 'junio': '06',
            'jul': '07', 'julio': '07',
            'ago': '08', 'aug': '08', 'agosto': '08',
            'sep': '09', 'sept': '09', 'septiembre': '09',
            'oct': '10', 'octubre': '10',
            'nov': '11', 'noviembre': '11',
            'dic': '12', 'dec': '12', 'diciembre': '12'
        }
        
        parsed = []
        for date_str in date_strings:
            try:
                parts = str(date_str).lower().split('-')
                if len(parts) >= 2:
                    month = months.get(parts[0], '01')
                    year = parts[1]
                    if len(year) == 2:
                        year = '20' + year
                    parsed.append(f"{year}-{month}-01")
                else:
                    parsed.append(None)
            except:
                parsed.append(None)
        return parsed
    
    def create_charts(self):
        """Generate all charts"""
        self.chart_files = []
        
        print("   Creating Production Overview...")
        self._create_production_charts()
        
        print("   Creating Efficiency Analysis...")
        self._create_efficiency_charts()
        
        print("   Creating Comparative Analysis...")
        self._create_comparative_charts()
        
        print("   Creating Trend Analysis...")
        self._create_trend_charts()
        
    def _create_production_charts(self):
        """Create production charts"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Production Overview', fontsize=16, fontweight='bold')
        
        # Chart 1: Ore Production
        ax1 = axes[0, 0]
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            ax1.plot(self.clean_df.index, self.clean_df['Ore_Mined_RGM'], 
                    marker='o', label='RGM', linewidth=2, markersize=4, color='#3498db')
        if 'Ore_Mined_Sar' in self.clean_df.columns:
            ax1.plot(self.clean_df.index, self.clean_df['Ore_Mined_Sar'], 
                    marker='s', label='Sar', linewidth=2, markersize=4, color='#e74c3c')
        ax1.set_title('Ore Production Over Time')
        ax1.set_xlabel('Date')
        ax1.set_ylabel('Ore Mined (kt)')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 2: Overburden
        ax2 = axes[0, 1]
        if 'Overburden_RGM' in self.clean_df.columns:
            ax2.plot(self.clean_df.index, self.clean_df['Overburden_RGM'], 
                    marker='o', label='RGM', linewidth=2, markersize=4, color='#9b59b6')
        if 'Overburden_Sar' in self.clean_df.columns:
            ax2.plot(self.clean_df.index, self.clean_df['Overburden_Sar'], 
                    marker='s', label='Sar', linewidth=2, markersize=4, color='#f39c12')
        ax2.set_title('Overburden Movement')
        ax2.set_xlabel('Date')
        ax2.set_ylabel('Overburden (kt)')
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 3: Total Material
        ax3 = axes[1, 0]
        total_ore = pd.Series(0, index=self.clean_df.index)
        total_overburden = pd.Series(0, index=self.clean_df.index)
        
        for col in self.clean_df.columns:
            if 'Ore_Mined' in col:
                total_ore += self.clean_df[col]
            elif 'Overburden' in col:
                total_overburden += self.clean_df[col]
        
        ax3.bar(self.clean_df.index, total_ore, label='Total Ore', alpha=0.7, color='#3498db')
        ax3.bar(self.clean_df.index, total_overburden, bottom=total_ore, 
               label='Total Overburden', alpha=0.7, color='#e67e22')
        ax3.set_title('Total Material Movement')
        ax3.set_xlabel('Date')
        ax3.set_ylabel('Material (kt)')
        ax3.legend()
        ax3.grid(True, alpha=0.3, axis='y')
        plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 4: Strip Ratio
        ax4 = axes[1, 1]
        if 'Overburden_RGM' in self.clean_df.columns and 'Ore_Mined_RGM' in self.clean_df.columns:
            strip_rgm = self.clean_df['Overburden_RGM'] / (self.clean_df['Ore_Mined_RGM'] + 0.001)
            strip_rgm = strip_rgm.replace([np.inf, -np.inf], np.nan)
            ax4.plot(self.clean_df.index, strip_rgm, marker='o', label='RGM', linewidth=2, markersize=4)
        if 'Overburden_Sar' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            strip_sar = self.clean_df['Overburden_Sar'] / (self.clean_df['Ore_Mined_Sar'] + 0.001)
            strip_sar = strip_sar.replace([np.inf, -np.inf], np.nan)
            ax4.plot(self.clean_df.index, strip_sar, marker='s', label='Sar', linewidth=2, markersize=4)
        ax4.set_title('Stripping Ratio Trends')
        ax4.set_xlabel('Date')
        ax4.set_ylabel('Strip Ratio')
        ax4.legend()
        ax4.grid(True, alpha=0.3)
        plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45)
        
        plt.tight_layout()
        filename = os.path.join(self.temp_dir, 'production.png')
        plt.savefig(filename, dpi=100, bbox_inches='tight')
        plt.close()
        self.chart_files.append(('Production Overview', filename))
        
    def _create_efficiency_charts(self):
        """Create efficiency charts"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Efficiency Analysis', fontsize=16, fontweight='bold')
        
        # Chart 1: Fleet vs Production
        ax1 = axes[0, 0]
        if 'Active_Fleet_Count_Aprox' in self.clean_df.columns:
            total_prod = pd.Series(0, index=self.clean_df.index)
            for col in self.clean_df.columns:
                if 'Ore_Mined' in col:
                    total_prod += self.clean_df[col]
            
            ax1_twin = ax1.twinx()
            ax1.bar(self.clean_df.index, total_prod, alpha=0.5, color='skyblue', label='Production')
            ax1_twin.plot(self.clean_df.index, self.clean_df['Active_Fleet_Count_Aprox'], 
                         color='red', marker='o', linewidth=2, markersize=4, label='Fleet')
            ax1.set_xlabel('Date')
            ax1.set_ylabel('Production (kt)', color='blue')
            ax1_twin.set_ylabel('Fleet Count', color='red')
            ax1.set_title('Fleet vs Production')
            ax1.grid(True, alpha=0.3)
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 2: Diesel Consumption
        ax2 = axes[0, 1]
        if 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
            diesel = self.clean_df['Liter_of_Diesel_Consumed'] / 1000000
            ax2.plot(self.clean_df.index, diesel, marker='o', color='green', linewidth=2, markersize=4)
            ax2.fill_between(self.clean_df.index, diesel, alpha=0.3, color='green')
            ax2.set_title('Diesel Consumption')
            ax2.set_xlabel('Date')
            ax2.set_ylabel('Diesel (Million L)')
            ax2.grid(True, alpha=0.3)
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 3: Productivity
        ax3 = axes[1, 0]
        if 'Active_Fleet_Count_Aprox' in self.clean_df.columns:
            total_mat = pd.Series(0, index=self.clean_df.index)
            for col in self.clean_df.columns:
                if 'Ore_Mined' in col or 'Overburden' in col:
                    total_mat += self.clean_df[col]
            productivity = total_mat / (self.clean_df['Active_Fleet_Count_Aprox'] + 0.001)
            ax3.plot(self.clean_df.index, productivity, marker='o', color='purple', linewidth=2, markersize=4)
            ax3.set_title('Productivity per Fleet Unit')
            ax3.set_xlabel('Date')
            ax3.set_ylabel('kt per Unit')
            ax3.grid(True, alpha=0.3)
            plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 4: Fuel Efficiency
        ax4 = axes[1, 1]
        if 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
            total_mat = pd.Series(0, index=self.clean_df.index)
            for col in self.clean_df.columns:
                if 'Ore_Mined' in col or 'Overburden' in col:
                    total_mat += self.clean_df[col]
            fuel_eff = self.clean_df['Liter_of_Diesel_Consumed'] / (total_mat + 0.001)
            fuel_eff = fuel_eff.replace([np.inf, -np.inf], np.nan)
            ax4.plot(self.clean_df.index, fuel_eff, marker='o', color='orange', linewidth=2, markersize=4)
            ax4.set_title('Fuel Efficiency')
            ax4.set_xlabel('Date')
            ax4.set_ylabel('L/kt')
            ax4.grid(True, alpha=0.3)
            plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45)
        
        plt.tight_layout()
        filename = os.path.join(self.temp_dir, 'efficiency.png')
        plt.savefig(filename, dpi=100, bbox_inches='tight')
        plt.close()
        self.chart_files.append(('Efficiency Analysis', filename))
        
    def _create_comparative_charts(self):
        """Create comparative charts"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Comparative Analysis: RGM vs Sar', fontsize=16, fontweight='bold')
        
        # Chart 1: Production Pie
        ax1 = axes[0, 0]
        if 'Ore_Mined_RGM' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            rgm = self.clean_df['Ore_Mined_RGM'].sum()
            sar = self.clean_df['Ore_Mined_Sar'].sum()
            if rgm > 0 or sar > 0:
                ax1.pie([rgm, sar], labels=['RGM', 'Sar'], autopct='%1.1f%%', 
                       colors=['#3498db', '#e74c3c'])
                ax1.set_title('Ore Production Share')
        
        # Chart 2: Monthly Comparison
        ax2 = axes[0, 1]
        if 'Ore_Mined_RGM' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            sample = self.clean_df.iloc[::3] if len(self.clean_df) > 15 else self.clean_df
            x = np.arange(len(sample))
            width = 0.35
            ax2.bar(x - width/2, sample['Ore_Mined_RGM'], width, label='RGM', color='#3498db')
            ax2.bar(x + width/2, sample['Ore_Mined_Sar'], width, label='Sar', color='#e74c3c')
            ax2.set_xlabel('Date')
            ax2.set_ylabel('Ore (kt)')
            ax2.set_title('Monthly Comparison')
            ax2.set_xticks(x)
            ax2.set_xticklabels([d.strftime('%b-%y') for d in sample.index], rotation=45, ha='right')
            ax2.legend()
            ax2.grid(True, alpha=0.3, axis='y')
        
        # Chart 3: Overburden Pie
        ax3 = axes[1, 0]
        if 'Overburden_RGM' in self.clean_df.columns and 'Overburden_Sar' in self.clean_df.columns:
            rgm_ob = self.clean_df['Overburden_RGM'].sum()
            sar_ob = self.clean_df['Overburden_Sar'].sum()
            if rgm_ob > 0 or sar_ob > 0:
                ax3.pie([rgm_ob, sar_ob], labels=['RGM', 'Sar'], autopct='%1.1f%%',
                       colors=['#9b59b6', '#f39c12'])
                ax3.set_title('Overburden Share')
        
        # Chart 4: Metrics
        ax4 = axes[1, 1]
        if 'Ore_Mined_RGM' in self.clean_df.columns and 'Ore_Mined_Sar' in self.clean_df.columns:
            metrics = ['Avg\n(kt/mo)', 'Max\n(kt)', 'Total\n(kt/1000)']
            rgm_vals = [
                self.clean_df['Ore_Mined_RGM'].mean(),
                self.clean_df['Ore_Mined_RGM'].max(),
                self.clean_df['Ore_Mined_RGM'].sum()/1000
            ]
            sar_vals = [
                self.clean_df['Ore_Mined_Sar'].mean(),
                self.clean_df['Ore_Mined_Sar'].max(),
                self.clean_df['Ore_Mined_Sar'].sum()/1000
            ]
            x = np.arange(len(metrics))
            width = 0.35
            ax4.bar(x - width/2, rgm_vals, width, label='RGM', color='#3498db')
            ax4.bar(x + width/2, sar_vals, width, label='Sar', color='#e74c3c')
            ax4.set_ylabel('Values')
            ax4.set_title('Performance Metrics')
            ax4.set_xticks(x)
            ax4.set_xticklabels(metrics)
            ax4.legend()
            ax4.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        filename = os.path.join(self.temp_dir, 'comparative.png')
        plt.savefig(filename, dpi=100, bbox_inches='tight')
        plt.close()
        self.chart_files.append(('Comparative Analysis', filename))
        
    def _create_trend_charts(self):
        """Create trend charts"""
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle('Trend Analysis', fontsize=16, fontweight='bold')
        
        # Chart 1: Moving Averages
        ax1 = axes[0, 0]
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            ma3 = self.clean_df['Ore_Mined_RGM'].rolling(window=3, center=True).mean()
            ma6 = self.clean_df['Ore_Mined_RGM'].rolling(window=6, center=True).mean()
            ax1.plot(self.clean_df.index, self.clean_df['Ore_Mined_RGM'], alpha=0.3, label='Actual', color='gray')
            ax1.plot(self.clean_df.index, ma3, label='3-Month MA', linewidth=2, color='red')
            ax1.plot(self.clean_df.index, ma6, label='6-Month MA', linewidth=2, color='green')
            ax1.set_title('Moving Averages - RGM')
            ax1.set_xlabel('Date')
            ax1.set_ylabel('Ore (kt)')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
        
        # Chart 2: Yearly
        ax2 = axes[0, 1]
        yearly = {}
        for col in ['Ore_Mined_RGM', 'Ore_Mined_Sar']:
            if col in self.clean_df.columns:
                yr = self.clean_df[col].groupby(self.clean_df.index.year).sum()
                for y, v in yr.items():
                    if y not in yearly:
                        yearly[y] = {}
                    yearly[y][col] = v
        
        if yearly:
            years = sorted(yearly.keys())
            rgm_y = [yearly[y].get('Ore_Mined_RGM', 0) for y in years]
            sar_y = [yearly[y].get('Ore_Mined_Sar', 0) for y in years]
            x = np.arange(len(years))
            width = 0.35
            ax2.bar(x - width/2, rgm_y, width, label='RGM', color='#3498db')
            ax2.bar(x + width/2, sar_y, width, label='Sar', color='#e74c3c')
            ax2.set_xlabel('Year')
            ax2.set_ylabel('Total Ore (kt)')
            ax2.set_title('Yearly Production')
            ax2.set_xticks(x)
            ax2.set_xticklabels(years)
            ax2.legend()
            ax2.grid(True, alpha=0.3, axis='y')
        
        # Chart 3: Correlation
        ax3 = axes[1, 0]
        if 'Active_Fleet_Count_Aprox' in self.clean_df.columns and 'Liter_of_Diesel_Consumed' in self.clean_df.columns:
            fleet = self.clean_df['Active_Fleet_Count_Aprox']
            diesel = self.clean_df['Liter_of_Diesel_Consumed'] / 1000000
            mask = (fleet > 0) & (diesel > 0)
            if mask.sum() > 0:
                ax3.scatter(fleet[mask], diesel[mask], alpha=0.6, s=50)
                z = np.polyfit(fleet[mask], diesel[mask], 1)
                p = np.poly1d(z)
                ax3.plot(sorted(fleet[mask]), p(sorted(fleet[mask])), "r--", alpha=0.8)
                ax3.set_xlabel('Fleet Count')
                ax3.set_ylabel('Diesel (Million L)')
                ax3.set_title('Fleet vs Diesel')
                ax3.grid(True, alpha=0.3)
        
        # Chart 4: Seasonality
        ax4 = axes[1, 1]
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            monthly = self.clean_df['Ore_Mined_RGM'].groupby(self.clean_df.index.month).mean()
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            ax4.bar(range(1, 13), monthly.reindex(range(1, 13), fill_value=0), color='steelblue', alpha=0.7)
            avg = monthly.mean()
            ax4.axhline(y=avg, color='red', linestyle='--', label=f'Avg: {avg:.1f}')
            ax4.set_xlabel('Month')
            ax4.set_ylabel('Avg Ore (kt)')
            ax4.set_title('Monthly Seasonality - RGM')
            ax4.set_xticks(range(1, 13))
            ax4.set_xticklabels(months, rotation=45, ha='right')
            ax4.legend()
            ax4.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        filename = os.path.join(self.temp_dir, 'trends.png')
        plt.savefig(filename, dpi=100, bbox_inches='tight')
        plt.close()
        self.chart_files.append(('Trend Analysis', filename))
        
    def create_excel_with_charts(self):
        """Create Excel file with embedded charts"""
        if not self.output_file:
            base = os.path.splitext(os.path.basename(self.input_file))[0]
            self.output_file = f"{base}_with_charts.xlsx"
        elif not self.output_file.endswith('.xlsx'):
            self.output_file += '.xlsx'
        
        # Create workbook
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._add_summary(ws_summary)
        
        # Data sheet
        ws_data = wb.create_sheet("Data")
        for r in dataframe_to_rows(self.clean_df.reset_index(), index=False, header=True):
            ws_data.append(r)
        self._format_sheet(ws_data)
        
        # Chart sheets
        for name, file in self.chart_files:
            ws = wb.create_sheet(title=name)
            ws['A1'] = name
            ws['A1'].font = Font(size=16, bold=True)
            img = OpenpyxlImage(file)
            img.width = 1100
            img.height = 750
            ws.add_image(img, 'A3')
        
        # Save
        wb.save(self.output_file)
        return self.output_file
        
    def _add_summary(self, ws):
        """Add summary statistics"""
        data = [
            ['Mining Data Analysis Summary'],
            [''],
            ['Period', f"{self.clean_df.index[0].strftime('%b %Y')} to {self.clean_df.index[-1].strftime('%b %Y')}"],
            ['Months', str(len(self.clean_df))],
            [''],
            ['Metric', 'RGM', 'Sar'],
        ]
        
        # Add metrics
        if 'Ore_Mined_RGM' in self.clean_df.columns:
            rgm_total = self.clean_df['Ore_Mined_RGM'].sum()
            rgm_avg = self.clean_df['Ore_Mined_RGM'].mean()
        else:
            rgm_total = rgm_avg = 0
            
        if 'Ore_Mined_Sar' in self.clean_df.columns:
            sar_total = self.clean_df['Ore_Mined_Sar'].sum()
            sar_avg = self.clean_df['Ore_Mined_Sar'].mean()
        else:
            sar_total = sar_avg = 0
        
        data.append(['Total Ore (kt)', f'{rgm_total:,.0f}', f'{sar_total:,.0f}'])
        data.append(['Avg Ore/Month', f'{rgm_avg:,.1f}', f'{sar_avg:,.1f}'])
        
        for row in data:
            ws.append(row)
        
        # Format
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
    def _format_sheet(self, ws):
        """Format worksheet"""
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="366092")
            cell.alignment = Alignment(horizontal="center")


class MiningAnalyzerGUI:
    """GUI interface for the mining analyzer"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Mining Data Analyzer")
        self.root.geometry("600x550")
        self.processor = None
        self.setup_gui()
        
    def setup_gui(self):
        """Create GUI elements"""
        # Title
        tk.Label(self.root, text="Mining Data Analyzer", 
                font=('Arial', 18, 'bold')).pack(pady=20)
        
        # File selection
        frame1 = tk.LabelFrame(self.root, text="Step 1: Select File", padx=20, pady=20)
        frame1.pack(padx=20, pady=10, fill='x')
        
        self.file_label = tk.Label(frame1, text="No file selected", fg='gray')
        self.file_label.pack(pady=10)
        
        tk.Button(frame1, text="Browse Excel File", command=self.browse_file,
                 bg='#3498db', fg='white', padx=20, pady=10).pack()
        
        # Options
        frame2 = tk.LabelFrame(self.root, text="Step 2: Options", padx=20, pady=20)
        frame2.pack(padx=20, pady=10, fill='x')
        
        tk.Label(frame2, text="Sheet (optional):").pack()
        self.sheet_entry = tk.Entry(frame2)
        self.sheet_entry.pack(pady=5)
        
        tk.Label(frame2, text="Output file (optional):").pack()
        self.output_entry = tk.Entry(frame2)
        self.output_entry.pack(pady=5)
        
        # Generate
        frame3 = tk.LabelFrame(self.root, text="Step 3: Generate", padx=20, pady=20)
        frame3.pack(padx=20, pady=10, fill='x')
        
        self.generate_btn = tk.Button(frame3, text="Generate Charts", 
                                     command=self.generate,
                                     bg='#27ae60', fg='white', 
                                     padx=30, pady=15,
                                     font=('Arial', 12, 'bold'),
                                     state='disabled')
        self.generate_btn.pack(pady=10)
        
        self.status = tk.Label(frame3, text="Ready", fg='gray')
        self.status.pack()
        
    def browse_file(self):
        """Browse for file"""
        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")]
        )
        if filename:
            self.input_file = filename
            self.file_label.config(text=os.path.basename(filename), fg='black')
            self.generate_btn.config(state='normal')
            
    def generate(self):
        """Generate charts"""
        self.generate_btn.config(state='disabled')
        self.status.config(text="Processing...")
        
        def process():
            try:
                sheet = self.sheet_entry.get().strip() or None
                output = self.output_entry.get().strip() or None
                
                processor = MiningDataProcessor(self.input_file, sheet, output)
                output_file = processor.process()
                
                self.root.after(0, lambda: self.complete(output_file))
            except Exception as e:
                self.root.after(0, lambda: self.error(str(e)))
        
        thread = threading.Thread(target=process)
        thread.start()
        
    def complete(self, output_file):
        """Handle completion"""
        self.generate_btn.config(state='normal')
        self.status.config(text="✅ Complete!")
        messagebox.showinfo("Success", f"Charts created!\n\nFile: {output_file}")
        
    def error(self, msg):
        """Handle error"""
        self.generate_btn.config(state='normal')
        self.status.config(text="❌ Error")
        messagebox.showerror("Error", f"Error: {msg}")
        
    def run(self):
        """Run GUI"""
        self.root.mainloop()


def main():
    """Main entry point"""
    print("=" * 60)
    print("  Mining Data Analyzer - Complete Solution")
    print("=" * 60)
    
    if len(sys.argv) > 1:
        # Command-line mode
        input_file = sys.argv[1]
        sheet = sys.argv[2] if len(sys.argv) > 2 else None
        output = sys.argv[3] if len(sys.argv) > 3 else None
        
        if not os.path.exists(input_file):
            print(f"Error: File not found: {input_file}")
            sys.exit(1)
        
        processor = MiningDataProcessor(input_file, sheet, output)
        processor.process()
        
    elif GUI_AVAILABLE:
        # GUI mode
        print("\nStarting GUI mode...")
        app = MiningAnalyzerGUI()
        app.run()
        
    else:
        # Interactive command-line mode
        print("\nNo file specified. Interactive mode:")
        input_file = input("Enter Excel file path: ").strip()
        
        if not os.path.exists(input_file):
            print(f"Error: File not found: {input_file}")
            sys.exit(1)
        
        sheet = input("Enter sheet name (Enter for default): ").strip() or None
        output = input("Enter output filename (Enter for auto): ").strip() or None
        
        processor = MiningDataProcessor(input_file, sheet, output)
        processor.process()


if __name__ == "__main__":
    main()
